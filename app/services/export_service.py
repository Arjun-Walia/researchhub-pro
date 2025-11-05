"""Export service for generating reports in various formats."""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import io
import csv
import json
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.utils.exceptions import ExportError


logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting research data in multiple formats.
    """
    
    def __init__(self):
        """Initialize export service."""
        self.supported_formats = ['json', 'csv', 'markdown', 'html', 'txt']
        
        if REPORTLAB_AVAILABLE:
            self.supported_formats.append('pdf')
        
        if OPENPYXL_AVAILABLE:
            self.supported_formats.append('xlsx')
    
    def export_json(self, data: Any, pretty: bool = True) -> str:
        """
        Export data as JSON.
        
        Args:
            data: Data to export
            pretty: Pretty print JSON
            
        Returns:
            JSON string
        """
        try:
            if pretty:
                return json.dumps(data, indent=2, default=str)
            return json.dumps(data, default=str)
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            raise ExportError(f"JSON export failed: {str(e)}")
    
    def export_csv(self, results: List[Dict[str, Any]], fields: Optional[List[str]] = None) -> str:
        """
        Export search results as CSV.
        
        Args:
            results: List of result dictionaries
            fields: Fields to include (None = all fields)
            
        Returns:
            CSV string
        """
        try:
            if not results:
                return ""
            
            output = io.StringIO()
            
            # Determine fields
            if fields is None:
                fields = list(results[0].keys())
            
            writer = csv.DictWriter(output, fieldnames=fields, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(results)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise ExportError(f"CSV export failed: {str(e)}")
    
    def export_markdown(
        self,
        query: str,
        results: List[Dict[str, Any]],
        include_metadata: bool = True
    ) -> str:
        """
        Export as Markdown document.
        
        Args:
            query: Research query
            results: Search results
            include_metadata: Include metadata section
            
        Returns:
            Markdown string
        """
        try:
            md = f"# Research Results: {query}\n\n"
            
            if include_metadata:
                md += f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}  \n"
                md += f"**Total Results:** {len(results)}  \n\n"
            
            md += "---\n\n"
            
            for i, result in enumerate(results, 1):
                md += f"## {i}. {result.get('title', 'Untitled')}\n\n"
                md += f"**URL:** [{result.get('url', 'N/A')}]({result.get('url', '#')})\n\n"
                
                if result.get('author'):
                    md += f"**Author:** {result['author']}  \n"
                
                if result.get('published_date'):
                    md += f"**Published:** {result['published_date']}  \n"
                
                if result.get('snippet'):
                    md += f"\n{result['snippet']}\n\n"
                
                if result.get('key_points'):
                    md += "**Key Points:**\n"
                    for point in result['key_points']:
                        md += f"- {point}\n"
                    md += "\n"
                
                md += "---\n\n"
            
            return md
            
        except Exception as e:
            logger.error(f"Markdown export failed: {str(e)}")
            raise ExportError(f"Markdown export failed: {str(e)}")
    
    def export_html(
        self,
        query: str,
        results: List[Dict[str, Any]],
        template: str = "default"
    ) -> str:
        """
        Export as HTML document.
        
        Args:
            query: Research query
            results: Search results
            template: HTML template style
            
        Returns:
            HTML string
        """
        try:
            html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Research Results: {query}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            line-height: 1.6;
            color: #333;
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        .metadata {{
            background: #ecf0f1;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }}
        .result {{
            background: #fff;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 20px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .result h2 {{
            color: #2980b9;
            margin-top: 0;
        }}
        .result a {{
            color: #3498db;
            text-decoration: none;
        }}
        .result a:hover {{
            text-decoration: underline;
        }}
        .snippet {{
            color: #555;
            margin: 15px 0;
        }}
        .key-points {{
            background: #f8f9fa;
            padding: 15px;
            border-left: 4px solid #3498db;
            margin: 15px 0;
        }}
        .key-points ul {{
            margin: 0;
            padding-left: 20px;
        }}
    </style>
</head>
<body>
    <h1>Research Results: {query}</h1>
    
    <div class="metadata">
        <strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}<br>
        <strong>Total Results:</strong> {len(results)}
    </div>
"""
            
            for i, result in enumerate(results, 1):
                html += f"""
    <div class="result">
        <h2>{i}. {result.get('title', 'Untitled')}</h2>
        <p><strong>URL:</strong> <a href="{result.get('url', '#')}" target="_blank">{result.get('url', 'N/A')}</a></p>
"""
                
                if result.get('author'):
                    html += f"        <p><strong>Author:</strong> {result['author']}</p>\n"
                
                if result.get('published_date'):
                    html += f"        <p><strong>Published:</strong> {result['published_date']}</p>\n"
                
                if result.get('snippet'):
                    html += f"        <div class=\"snippet\">{result['snippet']}</div>\n"
                
                if result.get('key_points'):
                    html += "        <div class=\"key-points\">\n"
                    html += "            <strong>Key Points:</strong>\n"
                    html += "            <ul>\n"
                    for point in result['key_points']:
                        html += f"                <li>{point}</li>\n"
                    html += "            </ul>\n"
                    html += "        </div>\n"
                
                html += "    </div>\n"
            
            html += """
</body>
</html>"""
            
            return html
            
        except Exception as e:
            logger.error(f"HTML export failed: {str(e)}")
            raise ExportError(f"HTML export failed: {str(e)}")
    
    def export_txt(self, query: str, results: List[Dict[str, Any]]) -> str:
        """
        Export as plain text.
        
        Args:
            query: Research query
            results: Search results
            
        Returns:
            Plain text string
        """
        try:
            txt = f"RESEARCH RESULTS: {query}\n"
            txt += "=" * 80 + "\n\n"
            txt += f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}\n"
            txt += f"Total Results: {len(results)}\n\n"
            txt += "=" * 80 + "\n\n"
            
            for i, result in enumerate(results, 1):
                txt += f"{i}. {result.get('title', 'Untitled')}\n"
                txt += "-" * 80 + "\n"
                txt += f"URL: {result.get('url', 'N/A')}\n"
                
                if result.get('author'):
                    txt += f"Author: {result['author']}\n"
                
                if result.get('published_date'):
                    txt += f"Published: {result['published_date']}\n"
                
                txt += "\n"
                
                if result.get('snippet'):
                    txt += f"{result['snippet']}\n\n"
                
                if result.get('key_points'):
                    txt += "Key Points:\n"
                    for point in result['key_points']:
                        txt += f"  - {point}\n"
                    txt += "\n"
                
                txt += "=" * 80 + "\n\n"
            
            return txt
            
        except Exception as e:
            logger.error(f"Text export failed: {str(e)}")
            raise ExportError(f"Text export failed: {str(e)}")
    
    def export_xlsx(self, results: List[Dict[str, Any]], sheet_name: str = "Results") -> bytes:
        """
        Export as Excel file.
        
        Args:
            results: Search results
            sheet_name: Worksheet name
            
        Returns:
            Excel file bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ExportError("openpyxl not installed. Cannot export to Excel.")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not results:
                return self._workbook_to_bytes(wb)
            
            # Define fields
            fields = ['title', 'url', 'author', 'published_date', 'snippet', 'relevance_score']
            headers = ['Title', 'URL', 'Author', 'Published Date', 'Snippet', 'Relevance Score']
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            
            # Write data
            for row, result in enumerate(results, 2):
                for col, field in enumerate(fields, 1):
                    value = result.get(field, '')
                    ws.cell(row=row, column=col, value=str(value))
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 80
            ws.column_dimensions['F'].width = 15
            
            return self._workbook_to_bytes(wb)
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}")
    
    def _workbook_to_bytes(self, wb) -> bytes:
        """Convert workbook to bytes."""
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export(
        self,
        data: Dict[str, Any],
        format_type: str,
        **kwargs
    ) -> Any:
        """
        Export data in specified format.
        
        Args:
            data: Data to export (should include 'query' and 'results')
            format_type: Export format
            **kwargs: Additional format-specific parameters
            
        Returns:
            Exported data (string or bytes)
        """
        format_type = format_type.lower()
        
        if format_type not in self.supported_formats:
            raise ExportError(f"Unsupported format: {format_type}")
        
        query = data.get('query', 'Research')
        results = data.get('results', [])
        
        try:
            if format_type == 'json':
                return self.export_json(data, pretty=kwargs.get('pretty', True))
            elif format_type == 'csv':
                return self.export_csv(results, fields=kwargs.get('fields'))
            elif format_type == 'markdown':
                return self.export_markdown(query, results, include_metadata=kwargs.get('include_metadata', True))
            elif format_type == 'html':
                return self.export_html(query, results, template=kwargs.get('template', 'default'))
            elif format_type == 'txt':
                return self.export_txt(query, results)
            elif format_type == 'xlsx':
                return self.export_xlsx(results, sheet_name=kwargs.get('sheet_name', 'Results'))
            else:
                raise ExportError(f"Format handler not implemented: {format_type}")
                
        except Exception as e:
            logger.error(f"Export failed for format {format_type}: {str(e)}")
            raise ExportError(f"Export failed: {str(e)}")
